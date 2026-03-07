"""
Парсер XLSX формата.
"""

from typing import Any, Dict, List, Optional, Union
from pathlib import Path

from openpyxl import load_workbook, Workbook
from parser.base import BaseParser


class XlsxParser(BaseParser):
    """Парсер XLSX файлов."""

    extensions = [".xlsx", ".xlsm", ".xltx", ".xltm"]
    format_name = "xlsx"

    def parse(
        self,
        file_path: str,
        encoding: str = "utf-8",
        sheet: Optional[Union[str, int]] = None,
        has_header: bool = True,
        skip_empty_rows: bool = True,
        skip_empty_cols: bool = False,
        value_only: bool = True,
        **kwargs
    ) -> Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]]:
        """
        Парсинг XLSX файла.

        Args:
            file_path: Путь к файлу.
            encoding: Кодировка.
            sheet: Имя или индекс листа (None = все листы).
            has_header: Есть ли заголовок.
            skip_empty_rows: Пропускать пустые строки.
            skip_empty_cols: Пропускать пустые столбцы.
            value_only: Только значения (без формул).

        Returns:
            Список словарей или словарь по листам.
        """
        wb = load_workbook(file_path, data_only=value_only)

        if sheet is not None:
            # Один лист
            if isinstance(sheet, int):
                ws = wb.worksheets[sheet]
            else:
                ws = wb[sheet]

            data = self._parse_worksheet(ws, has_header, skip_empty_rows, skip_empty_cols)
            wb.close()
            return data
        else:
            # Все листы
            result = {}

            for ws in wb.worksheets:
                result[ws.title] = self._parse_worksheet(
                    ws, has_header, skip_empty_rows, skip_empty_cols
                )

            wb.close()
            return result

    def _parse_worksheet(
        self,
        ws,
        has_header: bool = True,
        skip_empty_rows: bool = True,
        skip_empty_cols: bool = False,
    ) -> List[Dict[str, Any]]:
        """Парсинг листа."""
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        # Пропуск пустых строк
        if skip_empty_rows:
            rows = [
                row for row in rows
                if any(cell is not None for cell in row)
            ]

        if not rows:
            return []

        # Пропуск пустых столбцов
        if skip_empty_cols:
            max_col = max(
                i for i, cell in enumerate(rows[0])
                if any(row[i] is not None for row in rows)
            ) + 1
            rows = [row[:max_col] for row in rows]

        # Заголовки
        if has_header and rows:
            headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
        else:
            headers = [f"column_{i}" for i in range(len(rows[0]))]
            data_rows = rows

        # Конвертация в словари
        result = []

        for row in data_rows:
            row_dict = {}

            for i, value in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    row_dict[key] = self._convert_value(value)
                else:
                    row_dict[f"column_{i}"] = self._convert_value(value)

            result.append(row_dict)

        return result

    def _convert_value(self, value: Any) -> Any:
        """Конвертация значения."""
        if value is None:
            return None

        # datetime объекты
        from datetime import datetime, date

        if isinstance(value, (datetime, date)):
            return value.isoformat()

        return value

    def save(
        self,
        data: Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
        file_path: str,
        encoding: str = "utf-8",
        sheet_name: str = "Sheet1",
        **kwargs
    ) -> None:
        """
        Сохранение XLSX файла.

        Args:
            data: Данные.
            file_path: Путь к файлу.
            encoding: Кодировка.
            sheet_name: Имя листа.
        """
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Удаляем стандартный лист
        default_sheet = wb.active

        if isinstance(data, dict):
            # Несколько листов
            for i, (name, sheet_data) in enumerate(data.items()):
                if i == 0:
                    ws = default_sheet
                    ws.title = name
                else:
                    ws = wb.create_sheet(name)

                self._write_worksheet(ws, sheet_data)
        else:
            # Один лист
            ws = default_sheet
            ws.title = sheet_name
            self._write_worksheet(ws, data)

        wb.save(file_path)
        wb.close()

    def _write_worksheet(self, ws, data: List[Dict[str, Any]]) -> None:
        """Запись данных в лист."""
        if not data:
            return

        # Заголовки
        headers = list(data[0].keys())

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        for row_idx, row in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header)
                ws.cell(row=row_idx, column=col_idx, value=value)

    def get_info(self, file_path: str) -> Dict[str, Any]:
        """
        Получение информации о файле.

        Args:
            file_path: Путь к файлу.

        Returns:
            Информация о файле.
        """
        wb = load_workbook(file_path, data_only=True)

        info = {
            "sheets": [],
            "total_sheets": len(wb.worksheets),
        }

        for ws in wb.worksheets:
            sheet_info = {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }

            # Подсчет непустых ячеек
            non_empty = sum(
                1 for row in ws.iter_rows(values_only=True)
                for cell in row if cell is not None
            )
            sheet_info["non_empty_cells"] = non_empty

            info["sheets"].append(sheet_info)

        wb.close()
        return info

    def to_csv(
        self,
        file_path: str,
        output_path: str,
        sheet: Optional[Union[str, int]] = None,
        delimiter: str = ",",
        **kwargs
    ) -> None:
        """
        Конвертация XLSX в CSV.

        Args:
            file_path: Путь к XLSX файлу.
            output_path: Путь к CSV файлу.
            sheet: Имя или индекс листа.
            delimiter: Разделитель.
        """
        import csv

        data = self.parse(file_path, sheet=sheet, has_header=True)

        if not data:
            return

        # Если список листов
        if isinstance(data, dict):
            for name, sheet_data in data.items():
                out_path = f"{output_path}_{name}.csv"
                self._write_csv(sheet_data, out_path, delimiter)
        else:
            self._write_csv(data, output_path, delimiter)

    def _write_csv(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        delimiter: str = ",",
    ) -> None:
        """Запись CSV файла."""
        import csv

        if not data:
            return

        headers = list(data[0].keys())

        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)

    def merge_sheets(
        self,
        file_path: str,
        output_path: Optional[str] = None,
        **kwargs
    ) -> List[Dict[str, Any]]:
        """
        Слияние всех листов в один.

        Args:
            file_path: Путь к файлу.
            output_path: Путь для сохранения.

        Returns:
            Слитые данные.
        """
        data = self.parse(file_path, sheet=None)

        if isinstance(data, dict):
            merged = []

            for sheet_data in data.values():
                merged.extend(sheet_data)
        else:
            merged = data

        if output_path:
            self.save(merged, output_path)

        return merged
